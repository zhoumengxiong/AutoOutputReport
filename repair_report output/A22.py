# -*- coding: utf-8 -*-
import pymysql
from openpyxl import load_workbook
import datetime
from openpyxl.styles import Border, Side, Font, Alignment


# 只列出常用配置，其他略
def style_range(ws, cell_range, border, fill, font, alignment):
    """
    Apply styles to a range of cells as if they were a single cell.
    :param alignment: An openpyxl alignment object
    :param ws:  Excel worksheet instance
    :param cell_range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def getNowYearWeek():
    # 当前时间年第几周的计算
    timenow = datetime.datetime.now()
    NowYearWeek = timenow.isocalendar()
    return NowYearWeek


def fillsheet(key1, value1):
    args = (key1, last_wk, model)
    args1 = (value1, last_wk1, model)
    cursor.execute(sql1, args)
    # 读取所有
    failure = cursor.fetchall()  # 形成一个二维tuple
    ws = wb[value1]
    # 计算当前周所在的列号
    col = getNowYearWeek()[1] - 1 - (
        int(ws.cell(row=20, column=4).value[2:]) - 4)
    # 坏机失效一次因匹配
    for i in range(len(failure)):
        for r in range(38, 64):
            if ws.cell(row=r, column=1).value == failure[i][0]:
                ws.cell(row=r, column=col).value = failure[i][1]
    # 执行SQL查询语句
    cursor.execute(sql2, args)
    act = cursor.fetchall()  # 形成一个二维tuple
    # 写入当周激活量
    ws.cell(row=35, column=col).value = act[0][0]
    cursor.execute(sql3, args1)
    sale = cursor.fetchall()  # 形成一个二维tuple
    # 写入前一周销量数据
    ws.cell(row=23, column=col - 1).value = sale[0][0]

    # 写入返修率
    sale_sum = 0
    fail_sum = 0
    act_sum = 0
    PCBA = 0
    LCD = 0
    TP = 0
    SW = 0
    for m in range(4, col + 1):
        # 计算total接机
        for r in range(38, 64):
            if isinstance(ws.cell(row=r, column=m).value, int):
                fail_sum = ws.cell(row=r, column=m).value + fail_sum
        if isinstance(ws.cell(row=23, column=m).value, int):  # 计算total销量
            sale_sum = ws.cell(row=23, column=m).value + sale_sum
        if isinstance(ws.cell(row=35, column=m).value, int):  # 计算total激活量
            act_sum = ws.cell(row=35, column=m).value + act_sum
        if isinstance(ws.cell(row=39, column=m).value, int):  # 计算total PCBA坏机数
            PCBA = ws.cell(row=39, column=m).value + PCBA
        if isinstance(ws.cell(row=44, column=m).value, int):  # 计算total LCD坏机数
            LCD = ws.cell(row=44, column=m).value + LCD
        if isinstance(ws.cell(row=41, column=m).value, int):  # 计算total TP坏机数
            TP = ws.cell(row=41, column=m).value + TP
        if isinstance(ws.cell(row=38, column=m).value, int):  # 计算total BAT坏机数
            SW = ws.cell(row=38, column=m).value + SW
        # 计算返修率
        ws.cell(row=22, column=m).value = ws.cell(row=64, column=m).value
        ws.cell(row=34, column=m).value = ws.cell(row=64, column=m).value
        if m == col - 1:
            ws.cell(
                row=21, column=m).value = fail_sum / (sale_sum + sale[0][0])
            ws.cell(row=30, column=m).value = fail_sum / act_sum
            ws.cell(row=31, column=m).value = PCBA / act_sum
            ws.cell(row=32, column=m).value = (LCD + TP) / act_sum
            ws.cell(row=33, column=m).value = SW / act_sum
        elif m == col:
            ws.cell(
                row=21, column=m).value = fail_sum / (sale_sum + sale[0][0])
            ws.cell(row=30, column=m).value = fail_sum / (act_sum + act[0][0])
            ws.cell(row=31, column=m).value = PCBA / (act_sum + act[0][0])
            ws.cell(
                row=32, column=m).value = (LCD + TP) / (act_sum + act[0][0])
            ws.cell(row=33, column=m).value = SW / (act_sum + act[0][0])
        else:
            ws.cell(row=21, column=m).value = fail_sum / sale_sum
            ws.cell(row=30, column=m).value = fail_sum / act_sum
            ws.cell(row=31, column=m).value = PCBA / act_sum
            ws.cell(row=32, column=m).value = (LCD + TP) / act_sum
            ws.cell(row=33, column=m).value = SW / act_sum

    # 清空P39:Q63范围值
    for row in ws.iter_rows(min_row=39, min_col=16, max_col=17, max_row=63):
        for cell in row:
            cell.value = ''

    # 写入当周的一次因透视表
    sum = 39
    for i in range(len(failure)):
        ws.cell(row=sum, column=16).value = failure[i][0]
        ws.cell(row=sum, column=17).value = failure[i][1]
        sum = sum + 1
    # 合并单元格样式设置
    cell_range = [
        'A20:A23', 'A25:A35', 'B20:C20', 'B21:C21', 'B22:C22', 'B23:C23',
        'B25:C25', 'B34:B35', 'P37:R37', 'B26:B29', 'B30:B33'
    ]
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    font = Font(b=False, color="000000", name="Segoe UI Semibold", size='9')
    al = Alignment(horizontal="center", vertical="center")
    for cell in cell_range:
        style_range(
            ws, cell, border=border, fill=None, font=font, alignment=al)


# 计算激活量和销量查询周数
last_wk = str(18) + str(getNowYearWeek()[1] - 1)
last_wk1 = str(18) + str(getNowYearWeek()[1] - 2)
model = "A22"
countries = {'India': '印度'}
config = {
    'host': "localhost",  # 本地的话就是这个
    'user': "Dream",  # 输入你的数据库账号
    'password': "Dream123$",  # 以及数据库密码
    'db': "raw_data",  # 数据库名（database名）
    'charset': 'utf8mb4'  # 读取中文不想乱码的话，记得设置这个
}
# 打开数据库连接
db = pymysql.connect(**config)  # 使用关键字参数特性，这样好看一些
# 使用 cursor() 方法创建一个游标对象 cursor
cursor = db.cursor()
# 创建SQL查询语句
sql1 = """select level1cause,count(*) as QTY from grandtotal where CountryName=%s and week=%s
    and model=%s group by Level1Cause order by QTY DESC"""
sql2 = """select sum(qty) from activation_qty where country=%s and week=%s and model=%s"""
sql3 = """select sum(shipqty) from sale where SaleCoutry=%s and week=%s and model2=%s"""
wb = load_workbook(
    r"/mnt/hgfs/Transsion/Repair rate report/repair_rate_report/transition/week23 A22 repair rate(warranty)1.xlsx"
)
for key, value in countries.items():
    fillsheet(key, value)
wb.save(
    filename=
    r"/mnt/hgfs/Transsion/Repair rate report/repair_rate_report/transition/week23 A22 repair rate(warranty)1.xlsx"
)
