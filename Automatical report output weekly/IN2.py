# -*- coding: utf-8 -*-
import pymysql
from openpyxl import load_workbook
import datetime


# 只列出常用配置，其他略
def getNowYearWeek():
    # 当前时间年第几周的计算
    timenow = datetime.datetime.now()
    NowYearWeek = timenow.isocalendar()
    return NowYearWeek


def fillsheet(key1, value1):
    args = (key1, last_wk, model)
    args1 = (value1, last_wk1, model)
    cursor.execute(sql1, args)
    # 读取所有
    failure = cursor.fetchall()  # 形成一个二维tuple
    ws = wb[value1]
    # 坏机失效一次因匹配
    for i in range(len(failure)):
        for r in range(40, 66):
            if ws.cell(row=r, column=1).value == failure[i][0]:
                ws.cell(row=r, column=col).value = failure[i][1]
    # 执行SQL查询语句
    cursor.execute(sql2, args)
    act = cursor.fetchall()  # 形成一个二维tuple
    # 写入当周激活量
    ws.cell(row=37, column=col).value = act[0][0]
    cursor.execute(sql3, args1)
    sale = cursor.fetchall()  # 形成一个二维tuple
    # 写入当周销量数据
    ws.cell(row=23, column=col - 1).value = sale[0][0]
    # 清空P41:Q65范围值
    for row in ws.iter_rows(min_row=41, min_col=16, max_col=17, max_row=65):
        for cell in row:
            cell.value = ''
    # 写入当周的一次因透视表
    sum = 41
    for i in range(len(failure)):
        ws.cell(row=sum, column=16).value = failure[i][0]
        ws.cell(row=sum, column=17).value = failure[i][1]
        sum = sum + 1


# 计算激活量和销量查询周数
last_wk = str(18) + str(getNowYearWeek()[1] - 1)
last_wk1 = str(18) + str(getNowYearWeek()[1] - 2)
# 计算当前周所在的列号
col = getNowYearWeek()[1] - 1 - 12
model = "IN2"
countries = {'India': '印度'}
config = {
    'host': "localhost",  # 本地的话就是这个
    'user': "Dream",  # 输入你的数据库账号
    'password': "Dream123$",  # 以及数据库密码
    'db': "raw_data",  # 数据库名（database名）
    'charset': 'utf8mb4'  # 读取中文不想乱码的话，记得设置这个
}
# 打开数据库连接
db = pymysql.connect(**config)  # 使用关键字参数特性，这样好看一些
# 使用 cursor() 方法创建一个游标对象 cursor
cursor = db.cursor()
# 创建SQL查询语句
sql1 = """select level1cause,count(*) as QTY from grandtotal where CountryName=%s and week=%s
    and model=%s group by Level1Cause order by QTY DESC"""
sql2 = """select sum(qty) from activation_qty where country=%s and week=%s and model=%s"""
sql3 = """select sum(shipqty) from sale where SaleCoutry=%s and week=%s and model2=%s"""
wb = load_workbook(r"C:\Users\Dream\week23 IN2 repair rate(warranty)1.xlsx")
for key, value in countries.items():
    fillsheet(key, value)
wb.save(filename=r"C:\Users\Dream\week23 IN2 repair rate(warranty)1.xlsx")
