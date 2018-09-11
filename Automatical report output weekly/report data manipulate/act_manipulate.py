# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
import datetime


def getNowYearWeek():
    # 当前时间年第几周的计算
    timenow = datetime.datetime.now()
    NowYearWeek = timenow.isocalendar()
    return NowYearWeek


def act_transform():
    wk = str(18) + str(getNowYearWeek()[1] - 1)
    # 此行引号下的路径改成你自己周激活量存放的路径，来源刘志奋
    wb = load_workbook(r"F:\share2linux\act.xlsx")
    for sheet in wb.sheetnames:
        if sheet != "激活量明细":
            print(sheet)
            # del wb["%s" % sheet]
            wb.remove(wb[sheet])
    df = pd.DataFrame(wb['激活量明细'].values)
    df = df.iloc[1:, :3]
    df.insert(3, 3, wk)
    df.loc[[1], [3]] = 'week'
    df.to_csv(r"F:\share2linux\act.csv", index=False, encoding='utf_8_sig',  # 此处为csv格式存放路径
              header=None)


if __name__ == '__main__':
    act_transform()
