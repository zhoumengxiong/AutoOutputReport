# -*- coding: utf-8 -*-
import pymysql
from openpyxl import load_workbook
import datetime
from openpyxl.styles import Border, Side, Font, Alignment


def style_range(ws, cell_range, border, fill, font, alignment):
    """
    Apply styles to a range of cells as if they were a single cell.
    :param alignment: An openpyxl alignment object
    :param ws:  Excel worksheet instance
    :param cell_range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def getNowYearWeek():
    # 当前时间年第几周的计算
    time_now = datetime.datetime.now()
    NowYearWeek = time_now.isocalendar()
    return NowYearWeek


def write_sale_act_fail(key1, value1):
    args = (key1, last_wk, model)
    args1 = (value1, last_wk1, model)
    # 查询当周坏机透视表
    cursor.execute(sql1, args)
    # 读取所有
    failure = cursor.fetchall()  # 形成一个二维tuple
    ws = wb[value1]
    # 坏机失效一次因匹配
    for i in range(len(failure)):
        for r in range(40, 66):
            if ws.cell(row=r, column=1).value == failure[i][0]:
                ws.cell(row=r, column=col).value = failure[i][1]
    # 执行SQL查询语句
    cursor.execute(sql2, args)
    act = cursor.fetchall()  # 形成一个二维tuple
    # 写入当周激活量
    ws.cell(row=37, column=col).value = act[0][0]
    cursor.execute(sql3, args1)
    sale = cursor.fetchall()  # 形成一个二维tuple
    # 写入前一周销量数据
    ws.cell(row=23, column=col - 1).value = sale[0][0]
    # 清空P41:Q65范围值
    for row in ws.iter_rows(min_row=41, min_col=16, max_col=17, max_row=65):
        for cell in row:
            cell.value = ''
    # 写入当周的一次因透视表
    sum = 41
    for i in range(len(failure)):
        ws.cell(row=sum, column=16).value = failure[i][0]
        ws.cell(row=sum, column=17).value = failure[i][1]
        sum = sum + 1


def write_repair_rate(value1):
    ws = wb[value1]
    # 计算并写入返修率
    sale_sum = 0
    fail_sum = 0
    act_sum = 0
    PCBA = 0
    LCD = 0
    TP = 0
    BAT = 0
    for m in range(4, col + 1):
        # 计算total接机
        for r in range(40, 66):
            if isinstance(ws.cell(row=r, column=m).value, int):
                fail_sum = ws.cell(row=r, column=m).value + fail_sum
        if isinstance(ws.cell(row=23, column=m).value, int):  # 计算total销量
            sale_sum = ws.cell(row=23, column=m).value + sale_sum
        if isinstance(ws.cell(row=37, column=m).value, int):  # 计算total激活量
            act_sum = ws.cell(row=37, column=m).value + act_sum
        if isinstance(ws.cell(row=41, column=m).value, int):  # 计算total PCBA坏机数
            PCBA = ws.cell(row=41, column=m).value + PCBA
        if isinstance(ws.cell(row=46, column=m).value, int):  # 计算total LCD坏机数
            LCD = ws.cell(row=46, column=m).value + LCD
        if isinstance(ws.cell(row=43, column=m).value, int):  # 计算total TP坏机数
            TP = ws.cell(row=43, column=m).value + TP
        if isinstance(ws.cell(row=45, column=m).value, int):  # 计算total BAT坏机数
            BAT = ws.cell(row=45, column=m).value + BAT
        # 计算返修率
        if sale_sum > 0:
            ws.cell(row=22, column=m).value = ws.cell(row=66, column=m).value
            if act_sum == 0:
                ws.cell(row=21, column=m).value = fail_sum / sale_sum
            else:
                ws.cell(row=36, column=m).value = ws.cell(row=66, column=m).value
                ws.cell(row=21, column=m).value = fail_sum / sale_sum
                ws.cell(row=31, column=m).value = fail_sum / act_sum
                ws.cell(row=32, column=m).value = PCBA / act_sum
                ws.cell(row=33, column=m).value = LCD / act_sum
                ws.cell(row=34, column=m).value = TP / act_sum
                ws.cell(row=35, column=m).value = BAT / act_sum
    # 合并单元格样式设置
    cells = ['A20:A23', 'A25:A37', 'B20:C20', 'B21:C21', 'B22:C22', 'B23:C23', 'B25:C25', 'B36:B37',
             'P39:R39', 'B26:B30', 'B31:B35']
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    font = Font(b=False, color="000000", name="Segoe UI Semibold", size='9')
    al = Alignment(horizontal="center", vertical="center")
    for area in cells:
        style_range(ws, area, border=border, fill=None, font=font, alignment=al)


# 计算激活量和销量查询周数
last_wk = str(18) + str(getNowYearWeek()[1] - 1)
last_wk1 = str(18) + str(getNowYearWeek()[1] - 2)
# 计算当前周所在的列号
# col=getNowYearWeek()[1]-1-16
model = "A32F"
countries = {'Nigeria': '尼日利亚', 'Senegal': '塞内加尔', 'Tanzania': '坦桑尼亚', 'Uganda': '乌干达', 'Cote dIvoire':
    '科特迪瓦', 'Zambia': '赞比亚', 'Ghana': '加纳', 'Kenya': '肯尼亚', 'Mali': '马里', 'Vietnam': '越南', 'Cameroon':
                 '喀麦隆', 'Bangladesh': '孟加拉', 'Egypt': '埃及'}
config = {
    'host': "localhost",  # 本地的话就是这个
    'user': "Dream",  # 输入你的数据库账号
    'password': "Dream123$",  # 以及数据库密码
    'db': "raw_data",  # 数据库名（database名）
    'charset': 'utf8mb4'  # 读取中文不想乱码的话，记得设置这个
}
# 打开数据库连接
db = pymysql.connect(**config)  # 使用关键字参数特性，这样好看一些
# 使用 cursor() 方法创建一个游标对象 cursor
cursor = db.cursor()
# 创建SQL查询语句
sql1 = """select level1cause,count(*) as QTY from grandtotal where CountryName=%s and week=%s
    and model=%s group by Level1Cause order by QTY DESC"""
sql2 = """select sum(qty) from activation_qty where country=%s and week=%s and model=%s"""
sql3 = """select sum(shipqty) from sale where SaleCoutry=%s and week=%s and model2=%s"""
path = r"F:\Transsion\Repair rate report\repair_rate_report\transition\week23 A32F repair rate(warranty)1.xlsx"
wb = load_workbook(filename=path)
# 计算当前周所在的列号
col = getNowYearWeek()[1] - 1 - (int(wb.active.cell(row=20, column=4).value[2:]) - 4)
# 向Workbook写入上周坏机匹配，激活量及上上周的销量
for key, value in countries.items():
    write_sale_act_fail(key, value)
wb.save(filename=path)
wb = load_workbook(filename=path)
for value in countries.values():
    write_repair_rate(value)
wb.save(filename=path)
