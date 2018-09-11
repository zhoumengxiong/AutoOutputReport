# -*- coding: utf-8 -*-
import pymysql
import openpyxl


def write_repair_rate(row1):
    # 计算并写入返修率
    fail_sum = 0
    act_sum = 0
    for col1 in range(2, len(months) + 2):
        try:
            cursor.execute(fail_qty, [models[row1 - 2], months[col1 - 2]])
            fail_sum = cursor.fetchall()[0][0] + fail_sum
            cursor.execute(act_qty, [models[row1 - 2], months[col1 - 2]])
            act_sum = cursor.fetchall()[0][0] + act_sum
            ws.cell(row=row, column=col1).value = fail_sum / act_sum
            wb.save(filename=path)
        except:
            continue


config = {
    'host': "localhost",  # 本地的话就是这个
    'user': "Dream",  # 输入你的数据库账号
    'password': "Dream123$",  # 以及数据库密码
    'db': "raw_data",  # 数据库名（database名）
    'charset': 'utf8mb4'  # 读取中文不想乱码的话，记得设置这个
}
# 打开数据库连接
db = pymysql.connect(**config)  # 使用关键字参数特性，这样好看一些
# 使用 cursor() 方法创建一个游标对象 cursor
cursor = db.cursor()
# 创建SQL查询语句
fail_qty = """select count(*) from call_month where model=%s and month=%s"""
act_qty = """select sum(qty) from act_month where model=%s and month=%s"""
models_qty = """select distinct model from call_month"""
month_qty = """select distinct mid(CallDate,1,7) from call_month"""
cursor.execute(models_qty)
models_tup = cursor.fetchall()
models = []
# 统计数据库里的机型清单
for x in range(len(models_tup)):
    models.append(models_tup[x][0])
# 统计数据库有记录的月份
months = []
cursor.execute(month_qty)
months_tup = cursor.fetchall()
for y in range(len(months_tup)):
    months.append(months_tup[y][0])
# 新建工作簿
wb = openpyxl.Workbook()
ws = wb.active
path = r'C:\Users\zhoum\Desktop\2018新机型返修率.xlsx'
columns = ['Model', '一月', '二月', '三月', '四月', '五月', '六月']
# 写入标题栏
for col in range(len(columns)):
    ws.cell(row=1, column=col + 1).value = columns[col]

for row in range(2, len(models) + 2):
    # 写入机型清单
    ws.cell(row=row, column=1).value = models[row - 2]
    write_repair_rate(row)
